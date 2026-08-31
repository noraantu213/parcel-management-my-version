"""
Script to convert all sheets in Sprint_2.xlsx to individual text files.
Each sheet is saved as a separate .txt file in the 'txt_output' folder.
"""

import openpyxl
import os
import json

XLSX_FILE = "Sprint_2.xlsx"
OUTPUT_DIR = "txt_output"

def convert_xlsx_to_txt(xlsx_path, output_dir):
    """Convert each sheet in the xlsx file to a separate txt file."""
    os.makedirs(output_dir, exist_ok=True)
    
    wb = openpyxl.load_workbook(xlsx_path)
    
    print(f"Found {len(wb.sheetnames)} sheet(s): {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        safe_name = sheet_name.replace("/", "_").replace("\\", "_").replace(" ", "_")
        txt_path = os.path.join(output_dir, f"{safe_name}.txt")
        
        with open(txt_path, "w", encoding="utf-8") as f:
            f.write(f"Sheet: {sheet_name}\n")
            f.write("=" * 80 + "\n\n")
            
            for row in ws.iter_rows(values_only=False):
                row_data = []
                for cell in row:
                    value = cell.value if cell.value is not None else ""
                    row_data.append(str(value))
                f.write("\t".join(row_data) + "\n")
        
        print(f"  Saved: {txt_path}")
    
    wb.close()
    print(f"\nAll sheets converted to text files in '{output_dir}' folder.")

if __name__ == "__main__":
    convert_xlsx_to_txt(XLSX_FILE, OUTPUT_DIR)
